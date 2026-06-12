from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
from database import get_db
import models, schemas
from auth import get_current_user, require_role
import openpyxl, io, math

router = APIRouter()

def get_elem(db, element_id):
    e = db.query(models.Element).filter(models.Element.id == element_id).first()
    if not e: raise HTTPException(status_code=404, detail="Élément introuvable")
    return e

# ── CRUD ─────────────────────────────────────────────────────────
@router.get("/", response_model=List[schemas.ElementOut])
def list_elements(
    projet_id: Optional[int] = None,
    lot: Optional[str] = None,
    type_element: Optional[str] = None,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user)
):
    q = db.query(models.Element)
    if projet_id:    q = q.filter(models.Element.projet_id == projet_id)
    if lot:          q = q.filter(models.Element.lot == lot)
    if type_element: q = q.filter(models.Element.type_element == type_element)
    return q.order_by(models.Element.reference).all()

@router.get("/{element_id}", response_model=schemas.ElementOut)
def get_element(element_id: int, db: Session = Depends(get_db),
                _: models.User = Depends(get_current_user)):
    return get_elem(db, element_id)

@router.post("/", response_model=schemas.ElementOut, status_code=201)
def create_element(data: schemas.ElementCreate, db: Session = Depends(get_db),
                   current_user: models.User = Depends(require_role("chef","suivi","topo","civil"))):
    e = models.Element(**data.model_dump(), createur_id=current_user.id)
    db.add(e); db.commit(); db.refresh(e)
    return e

@router.put("/{element_id}", response_model=schemas.ElementOut)
def update_element(element_id: int, data: schemas.ElementUpdate,
                   db: Session = Depends(get_db),
                   current_user: models.User = Depends(require_role("chef","suivi","topo","civil"))):
    e = get_elem(db, element_id)
    if e.statut_validation == models.StatutEnum.valide and current_user.role != models.RoleEnum.chef:
        raise HTTPException(status_code=403, detail="Élément validé — modification réservée au chef")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(e, k, v)
    db.commit(); db.refresh(e)
    return e

@router.delete("/{element_id}", status_code=204)
def delete_element(element_id: int, db: Session = Depends(get_db),
                   _: models.User = Depends(require_role("chef","suivi"))):
    e = get_elem(db, element_id)
    db.delete(e); db.commit()

# ── Workflow validation ──────────────────────────────────────────
@router.post("/{element_id}/soumettre", response_model=schemas.ElementOut)
def soumettre(element_id: int, db: Session = Depends(get_db),
              _: models.User = Depends(require_role("chef","suivi","topo","civil"))):
    e = get_elem(db, element_id)
    if e.statut_validation not in [models.StatutEnum.brouillon, models.StatutEnum.rejete]:
        raise HTTPException(status_code=400, detail="Seul un brouillon/rejeté peut être soumis")
    e.statut_validation = models.StatutEnum.en_revue
    db.commit(); db.refresh(e)
    return e

@router.post("/{element_id}/valider", response_model=schemas.ElementOut)
def valider(element_id: int, action: schemas.ValidationAction,
            db: Session = Depends(get_db),
            _: models.User = Depends(require_role("chef"))):
    e = get_elem(db, element_id)
    if e.statut_validation != models.StatutEnum.en_revue:
        raise HTTPException(status_code=400, detail="Seul un élément 'En revue' peut être validé/rejeté")
    if action.action == "valider":
        e.statut_validation = models.StatutEnum.valide
        e.commentaire_rejet = None
    elif action.action == "rejeter":
        if not action.commentaire:
            raise HTTPException(status_code=400, detail="Commentaire obligatoire pour rejeter")
        e.statut_validation = models.StatutEnum.rejete
        e.commentaire_rejet = action.commentaire
    else:
        raise HTTPException(status_code=400, detail="Action invalide")
    db.commit(); db.refresh(e)
    return e

# ── Import Excel ─────────────────────────────────────────────────
@router.post("/import/excel")
def import_excel(
    projet_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role("chef","suivi"))
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Fichier Excel requis")

    wb = openpyxl.load_workbook(io.BytesIO(file.file.read()), data_only=True)
    ws = wb.active

    # ── Détection automatique de la structure ─────────────────────
    row1 = [ws.cell(1, ci).value for ci in range(1, ws.max_column+1)]
    row2 = [ws.cell(2, ci).value for ci in range(1, ws.max_column+1)]

    is_warda_format = any(
        str(v).upper() in ("COORDONNEES VIROLE", "COORDONNEES THEORIQUES",
                           "PROFONDEURS (M)", "BETON ARME QTE (M3)")
        for v in row1 if v
    )

    if is_warda_format:
        col_names = []
        last_group = ""
        for r1, r2 in zip(row1, row2):
            if r1: last_group = str(r1).strip().upper().replace(" ","_")
            sub = str(r2).strip().upper() if r2 else ""
            col_names.append(f"{last_group}_{sub}" if sub else last_group)
        row3_val = ws.cell(3, 2).value
        start_row = 4 if (row3_val and not str(row3_val).startswith("PI")) else 3
    else:
        col_names = [str(v).strip().upper().replace(" ","_") if v else "" for v in row1]
        start_row = 2

    def get(d, *keys):
        for k in keys:
            v = d.get(k.upper())
            if v not in (None, "", "None", "/"):
                try: return float(v)
                except: pass
        return None

    def get_str(d, *keys):
        for k in keys:
            v = d.get(k.upper())
            if v and str(v).strip() not in ("", "None", "/"):
                return str(v).strip()
        return None

    def get_date(d, *keys):
        from datetime import datetime
        for k in keys:
            v = d.get(k.upper())
            if isinstance(v, datetime): return v
            if isinstance(v, str):
                for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                    try: return datetime.strptime(v, fmt)
                    except: pass
        return None

    created, updated, errors = 0, 0, []

    for row_values in ws.iter_rows(min_row=start_row, values_only=True):
        d = {col_names[i].upper(): v for i, v in enumerate(row_values) if i < len(col_names)}
        ref = get_str(d, "PIEUX", "REFERENCE", "REPERE")
        if not ref or str(ref).strip() == "/":
            continue
        try:
            theo_x   = get(d, "COORDONNEES_THEORIQUES_X", "COORD_X", "X")
            theo_y   = get(d, "COORDONNEES_THEORIQUES_Y", "COORD_Y", "Y")
            vir_x    = get(d, "COORDONNEES_VIROLE_X", "VIROLE_X")
            vir_y    = get(d, "COORDONNEES_VIROLE_Y", "VIROLE_Y")
            vir_z    = get(d, "COORDONNEES_VIROLE_Z", "VIROLE_Z")
            alt_tn   = get(d, "COORDONNEES_DU_TN_Z", "ALTITUDE_TN")
            cote_pl  = get(d, "COTE_PLANCHER_ZP", "COTE_PLANCHER", "ZP")
            cote_rec = get(d, "COTE_RECEPEE_A_ATTEINDRE", "COTE_RECEPEE")
            prof_toit= get(d, "PROFONDEURS_(M)_TOIT ROCHEUX", "PROFONDEURS_(M)_TOIT_ROCHEUX", "PROF_TOIT_ROCHEUX")
            prof_roche=get(d, "PROFONDEURS_(M)_ROCHE", "PROF_ROCHE")
            ancrage  = get(d, "ANCRAGE")
            if ancrage is None and prof_toit and prof_roche:
                ancrage = round(prof_roche - prof_toit, 4)
            vol = get(d, "BETON_ARME_QTE_(M3)", "VOLUME_FORE", "VOLUME_BETON")
            if vol is None and prof_roche:
                diam = get(d, "DIAMETRE") or 0.8
                vol  = round(math.pi * (diam/2)**2 * prof_roche, 4)
            sem_ref  = get_str(d, "SEMELLE")
            pot_ref  = get_str(d, "POTEAU")
            date_f   = get_date(d, "DATE", "DATE_FORAGE")
            statut   = get_str(d, "STATUT_ELEMENT") or "À faire"
            if vol and vol > 0 and statut == "À faire":
                statut = "Foré"

            fields = dict(
                projet_id=projet_id, reference=ref,
                lot=get_str(d, "LOT") or "Fondations",
                type_element=get_str(d, "TYPE_ELEMENT") or "Pieu",
                famille=get_str(d, "FAMILLE"),
                coord_x=theo_x, coord_y=theo_y,
                coord_virole_x=vir_x, coord_virole_y=vir_y, coord_virole_z=vir_z,
                altitude_tn=alt_tn,
                cote_plancher=cote_pl, cote_recepee=cote_rec,
                prof_toit_rocheux=prof_toit, prof_roche=prof_roche, ancrage=ancrage,
                volume_fore=vol, volume_budgetise=get(d, "VOLUME_BUDGETISE"),
                semelle_ref=sem_ref if sem_ref != "/" else None,
                poteau_ref=pot_ref if pot_ref != "/" else None,
                date_forage=date_f,
                materiau=get_str(d, "MATERIAU"),
                armature=get_str(d, "ARMATURE"),
                statut_element=statut,
            )

            e = db.query(models.Element).filter(
                models.Element.projet_id == projet_id,
                models.Element.reference == ref
            ).first()

            if e:
                for k, v in fields.items():
                    if v is not None: setattr(e, k, v)
                updated += 1
            else:
                clean = {k: v for k, v in fields.items() if v is not None}
                e = models.Element(**clean, createur_id=current_user.id)
                db.add(e)
                created += 1

        except Exception as ex:
            errors.append({"reference": ref, "erreur": str(ex)})

    db.commit()
    return {"créés": created, "mis_à_jour": updated, "erreurs": errors, "total": created + updated}


@router.get("/export/template-excel")
def export_template(_: models.User = Depends(get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pieux"
    headers = [
        "reference", "lot", "type_element", "famille",
        "coord_x", "coord_y",
        "coord_virole_x", "coord_virole_y", "coord_virole_z",
        "cote_plancher", "cote_recepee",
        "prof_toit_rocheux", "prof_roche", "ancrage",
        "diametre", "volume_budgetise", "volume_fore",
        "semelle_ref", "poteau_ref",
        "date_forage", "materiau", "armature", "statut_element"
    ]
    ws.append(headers)
    ws.append([
        "Pi.1", "Fondations", "Pieu", "F1",
        778930.920, 428785.296,
        778930.918, 428785.293, 715.95,
        715.26, 714.46,
        7.4, 8.52, 1.12,
        0.80, 0.603, 4.28,
        "/", "/",
        "2025-09-19", "Béton C30/37", "500A/B - Ø800", "Foré"
    ])
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_pieux.xlsx"}
    )

# ── Export DXF AutoCAD ───────────────────────────────────────────
@router.get("/export/dxf/{projet_id}")
def export_dxf(projet_id: int, db: Session = Depends(get_db),
               _: models.User = Depends(get_current_user)):
    try:
        import ezdxf
    except ImportError:
        raise HTTPException(status_code=500, detail="Module ezdxf non installé")

    elements = db.query(models.Element)\
                 .filter(models.Element.projet_id == projet_id).all()
    doc = ezdxf.new(dxfversion="R2010")
    doc.layers.new("PIEUX",    dxfattribs={"color": 1})
    doc.layers.new("SEMELLES", dxfattribs={"color": 3})
    doc.layers.new("POTEAUX",  dxfattribs={"color": 5})
    doc.layers.new("AUTRES",   dxfattribs={"color": 7})
    msp = doc.modelspace()
    LAYER_MAP = {"Pieu": "PIEUX", "Semelle": "SEMELLES", "Poteau": "POTEAUX"}

    for e in elements:
        x = e.coord_virole_x or e.coord_x
        y = e.coord_virole_y or e.coord_y
        if x is None or y is None:
            continue
        layer = LAYER_MAP.get(
            e.type_element.value if hasattr(e.type_element,'value') else e.type_element,
            "AUTRES"
        )
        r = (e.diametre or 0.8) / 2

        msp.add_circle(center=(x, y, 0), radius=r, dxfattribs={"layer": layer})
        msp.add_text(e.reference, dxfattribs={
            "layer": layer, "height": 0.15,
            "insert": (x + r + 0.1, y)
        })

        bs = e.coord_virole_z or e.cote_bs_reelle or e.cote_bs_theorique or 0
        prof = e.prof_roche or 1
        bi   = bs - prof
        msp.add_circle(center=(x, y, bs), radius=r, dxfattribs={"layer": layer})
        msp.add_circle(center=(x, y, bi), radius=r, dxfattribs={"layer": layer})
        for angle in [0, 90, 180, 270]:
            rad = math.radians(angle)
            px = x + r * math.cos(rad)
            py = y + r * math.sin(rad)
            msp.add_line((px, py, bs), (px, py, bi), dxfattribs={"layer": layer})

    buf = io.StringIO()
    doc.write(buf)
    content = buf.getvalue().encode('utf-8')
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/dxf",
        headers={"Content-Disposition": f"attachment; filename=smc_projet_{projet_id}.dxf"}
    )

# ── Stats ────────────────────────────────────────────────────────
@router.get("/stats/{projet_id}")
def stats(projet_id: int, db: Session = Depends(get_db),
          _: models.User = Depends(get_current_user)):
    elements = db.query(models.Element)\
                 .filter(models.Element.projet_id == projet_id).all()
    total = len(elements)
    if total == 0:
        return {"total": 0, "par_type": {}, "par_statut": {}, "volumes": {}}

    par_type = {}; par_statut = {}
    vol_budg = vol_fore = vol_recep = 0

    for e in elements:
        t = e.type_element.value if hasattr(e.type_element,'value') else str(e.type_element)
        s = e.statut_element.value if hasattr(e.statut_element,'value') else str(e.statut_element)
        par_type[t]   = par_type.get(t, 0) + 1
        par_statut[s] = par_statut.get(s, 0) + 1
        vol_budg  += e.volume_budgetise or 0
        vol_fore  += e.volume_fore or 0
        vol_recep += e.volume_recepé or 0

    return {
        "total": total,
        "par_type": par_type,
        "par_statut": par_statut,
        "volumes": {
            "budgetise": round(vol_budg, 3),
            "fore":      round(vol_fore, 3),
            "recepé":    round(vol_recep, 3),
            "delta":     round(vol_fore - vol_budg, 3),
        }
    }
